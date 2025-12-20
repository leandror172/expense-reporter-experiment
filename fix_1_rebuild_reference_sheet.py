"""
Script to rebuild "Referência de Categorias" sheet from actual expense sheet data.
This ensures the reference sheet perfectly matches the actual subcategories.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import OrderedDict

# Load workbook
print("Loading workbook...")
wb = openpyxl.load_workbook('Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2025.xlsx')

# Dictionary to store discovered subcategories
discovered = OrderedDict()

# Sheets to scan (skip Receitas and Listas de itens)
expense_sheets = {
    'Fixas': 'Fixas',
    'Variáveis': 'Variáveis',
    'Extras': 'Extras',
    'Adicionais': 'Adicionais'
}

print("\nScanning expense sheets for subcategories...")
print("=" * 80)

for sheet_type, sheet_name in expense_sheets.items():
    ws = wb[sheet_name]
    print(f"\n{sheet_name}:")

    current_category = None
    sheet_subcats = []

    # Scan column A (category) and column B (subcategory)
    for row in range(3, 300):  # Start from row 3, skip headers
        col_a = ws.cell(row, 1).value  # Category
        col_b = ws.cell(row, 2).value  # Subcategory

        # Update current category when found in column A
        if col_a and isinstance(col_a, str) and col_a.strip():
            current_category = col_a.strip()

        # Found a subcategory in column B
        if col_b and isinstance(col_b, str) and col_b.strip():
            subcat = col_b.strip()

            # Create unique key
            key = (sheet_type, current_category if current_category else "", subcat)

            if key not in discovered:
                discovered[key] = {
                    'sheet_type': sheet_type,
                    'category': current_category if current_category else "",
                    'subcategory': subcat,
                    'row': row
                }
                sheet_subcats.append(subcat)
                print(f"  Row {row}: {current_category or '(no category)'} > {subcat}")

print("\n" + "=" * 80)
print(f"Total subcategories found: {len(discovered)}")

# Add Receitas and Investimentos manually (from Listas de itens)
print("\nAdding Receitas and Investimentos from Listas de itens...")

receitas = [
    'Salário', 'Ajuda de custo', 'Aluguel', 'Pensão',
    'Horas extras', '13º salário', 'Férias', 'Outros'
]

investimentos = [
    'Ações', 'Tesouro Direto', 'Renda fixa',
    'Previdência privada', 'Outros'
]

# Add Receitas
for subcat in receitas:
    key = ('Receitas', '', subcat)
    if key not in discovered:
        discovered[key] = {
            'sheet_type': 'Receitas',
            'category': '',
            'subcategory': subcat,
            'row': 0
        }

# Add Investimentos
for subcat in investimentos:
    key = ('Investimentos', '', subcat)
    if key not in discovered:
        discovered[key] = {
            'sheet_type': 'Investimentos',
            'category': '',
            'subcategory': subcat,
            'row': 0
        }

print(f"Added {len(receitas)} Receitas subcategories")
print(f"Added {len(investimentos)} Investimentos subcategories")
print(f"\nTotal with Receitas/Investimentos: {len(discovered)}")

# Delete and recreate the reference sheet
print("\nRecreating 'Referência de Categorias' sheet...")

if 'Referência de Categorias' in wb.sheetnames:
    del wb['Referência de Categorias']

ws_ref = wb.create_sheet('Referência de Categorias')

# Set up formatting
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
category_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write title
ws_ref['A1'] = 'REFERÊNCIA DE CATEGORIAS E SUB-CATEGORIAS'
ws_ref['A1'].font = Font(bold=True, size=14)
ws_ref['A1'].alignment = Alignment(horizontal='center')
ws_ref.merge_cells('A1:D1')

# Add description
ws_ref['A2'] = 'Esta planilha lista todas as categorias e sub-categorias encontradas nas planilhas de despesas'
ws_ref['A2'].font = Font(italic=True, size=10)
ws_ref.merge_cells('A2:D2')

# Write headers
row = 4
ws_ref['A' + str(row)] = 'Tipo Principal'
ws_ref['B' + str(row)] = 'Categoria'
ws_ref['C' + str(row)] = 'Sub-categoria'
ws_ref['D' + str(row)] = 'Linha na Planilha'

for col in ['A', 'B', 'C', 'D']:
    cell = ws_ref[col + str(row)]
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

row += 1

# Write data organized by sheet type
current_sheet_type = None

for key, data in discovered.items():
    sheet_type = data['sheet_type']
    category = data['category']
    subcategory = data['subcategory']
    source_row = data['row']

    # Write data
    ws_ref.cell(row, 1).value = sheet_type
    ws_ref.cell(row, 2).value = category if category else ""
    ws_ref.cell(row, 3).value = subcategory
    ws_ref.cell(row, 4).value = source_row if source_row > 0 else ""

    # Apply formatting
    if sheet_type != current_sheet_type:
        current_sheet_type = sheet_type
        ws_ref.cell(row, 1).fill = category_fill
        ws_ref.cell(row, 1).font = category_font

    if category:
        ws_ref.cell(row, 2).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

    # Apply borders
    for col in range(1, 5):
        ws_ref.cell(row, col).border = thin_border

    # Align subcategory column left
    ws_ref.cell(row, 3).alignment = Alignment(horizontal='left', vertical='center')

    row += 1

# Adjust column widths
ws_ref.column_dimensions['A'].width = 20
ws_ref.column_dimensions['B'].width = 25
ws_ref.column_dimensions['C'].width = 30
ws_ref.column_dimensions['D'].width = 15

# Save workbook
print("Saving workbook...")
wb.save('Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2025.xlsx')

print("\n" + "=" * 80)
print("SUCCESS!")
print("=" * 80)
print(f"Reference sheet rebuilt with {row - 5} subcategories")
print("\nBreakdown by type:")

# Count by type
counts = {}
for key, data in discovered.items():
    sheet_type = data['sheet_type']
    counts[sheet_type] = counts.get(sheet_type, 0) + 1

for sheet_type, count in counts.items():
    print(f"  {sheet_type}: {count} subcategories")

print("\nThe reference sheet now matches actual expense sheets perfectly!")
