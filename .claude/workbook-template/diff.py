#!/usr/bin/env python3
"""Diff golden vs generated: inspect-dump (values/formulas/fills/bold/borders/merges/
widths/heights) AND openpyxl (font name/size/color/italic, numFmt, alignment, freeze)."""
import json, sys, os
from openpyxl import load_workbook

GOLD_DIR = ".claude/workbook-template/dump-reviewed"
GEN_DIR = ".claude/workbook-template/dump-generated"
GOLD_X = ".claude/workbook-template/template-reviewed.xlsx"
GEN_X = ".claude/workbook-template/template.xlsx"

SHEETS = ["Fixas", "Variáveis", "Receitas", "Adicionais", "Extras", "Listas de itens"]

def load(d, s):
    p = os.path.join(d, s + ".json")
    return json.load(open(p)) if os.path.exists(p) else None

def cellmap(sheet):
    """col-row -> cell dict from inspect rows."""
    m = {}
    for r in sheet.get("rows", []):
        for c in (r.get("cells") or []):
            m[f"{c['col']}{r['row']}"] = c
    return m

def diff_inspect(out):
    for s in SHEETS:
        g = load(GOLD_DIR, s); n = load(GEN_DIR, s)
        if g is None or n is None:
            out.append(f"[{s}] MISSING dump g={g is not None} n={n is not None}"); continue
        if g["dimensions"] != n["dimensions"]:
            out.append(f"[{s}] dims {g['dimensions']} != {n['dimensions']}")
        # merges
        gm = {m['range']: m.get('value','') for m in g['mergedCells']}
        nm = {m['range']: m.get('value','') for m in n['mergedCells']}
        for r in sorted(set(gm)|set(nm)):
            if r not in nm: out.append(f"[{s}] merge MISSING {r}={gm[r]!r}")
            elif r not in gm: out.append(f"[{s}] merge EXTRA {r}={nm[r]!r}")
            elif gm[r]!=nm[r]: out.append(f"[{s}] merge {r} val {gm[r]!r}!={nm[r]!r}")
        # widths/heights
        for k in sorted(set(g['columnWidths'])|set(n['columnWidths'])):
            gv=g['columnWidths'].get(k); nv=n['columnWidths'].get(k)
            if gv is None or nv is None or abs(gv-nv)>0.5:
                out.append(f"[{s}] width {k} {gv}!={nv}")
        for k in sorted(set(g['rowHeights'])|set(n['rowHeights']), key=lambda x:int(x)):
            gv=g['rowHeights'].get(k); nv=n['rowHeights'].get(k)
            if gv is None or nv is None or abs(gv-nv)>0.3:
                out.append(f"[{s}] height {k} {gv}!={nv}")
        # cells
        gc=cellmap(g); nc=cellmap(n)
        for k in sorted(set(gc)|set(nc), key=lambda x:(x[0].rjust(3),int(''.join(filter(str.isdigit,x))))):
            gv=gc.get(k); nv=nc.get(k)
            if gv is None:
                if nv.get('value','')!='' or nv.get('formula'): out.append(f"[{s}] {k} EXTRA v={nv.get('value')!r} f={nv.get('formula')!r}")
                continue
            if nv is None:
                if gv.get('value','')!='' or gv.get('formula'): out.append(f"[{s}] {k} MISSING v={gv.get('value')!r} f={gv.get('formula')!r}")
                continue
            # Suppress cached-value-only diffs: golden has a formula + cached result,
            # generated has the same formula but no cached value (populated on recalc).
            cached_only = gv.get('formula') and gv.get('formula')==nv.get('formula') and (nv.get('value','') in ('','0'))
            if str(gv.get('value',''))!=str(nv.get('value','')) and not cached_only:
                out.append(f"[{s}] {k} value {gv.get('value')!r}!={nv.get('value')!r}")
            if gv.get('formula')!=nv.get('formula'): out.append(f"[{s}] {k} formula {gv.get('formula')!r}!={nv.get('formula')!r}")
            gs=gv.get('style',{}); ns=nv.get('style',{})
            for attr in ['bgColor','bold','borderTop','borderBottom','borderLeft','borderRight']:
                if gs.get(attr)!=ns.get(attr): out.append(f"[{s}] {k} style.{attr} {gs.get(attr)!r}!={ns.get(attr)!r}")

def fcolor(c):
    if c is None or c.rgb is None: return None
    return str(c.rgb)[-6:] if isinstance(c.rgb,str) else None

def diff_openpyxl(out):
    wg=load_workbook(GOLD_X); wn=load_workbook(GEN_X)
    for s in SHEETS:
        if s not in wg.sheetnames or s not in wn.sheetnames:
            out.append(f"[op {s}] sheet missing"); continue
        sg=wg[s]; sn=wn[s]
        if sg.freeze_panes!=sn.freeze_panes:
            out.append(f"[op {s}] freeze {sg.freeze_panes}!={sn.freeze_panes}")
        maxr=max(sg.max_row,sn.max_row); maxc=max(sg.max_column,sn.max_column)
        for row in range(1,maxr+1):
            for col in range(1,maxc+1):
                cg=sg.cell(row=row,column=col); cn=sn.cell(row=row,column=col)
                ref=cg.coordinate
                # only check cells that have content or style on either side
                if cg.value is None and cn.value is None and cg.font.name is None and cn.font.name is None:
                    pass
                if cg.font.name!=cn.font.name: out.append(f"[op {s}] {ref} font.name {cg.font.name}!={cn.font.name}")
                if (cg.font.size or 0)!=(cn.font.size or 0): out.append(f"[op {s}] {ref} font.size {cg.font.size}!={cn.font.size}")
                if bool(cg.font.bold)!=bool(cn.font.bold): out.append(f"[op {s}] {ref} bold {cg.font.bold}!={cn.font.bold}")
                if bool(cg.font.italic)!=bool(cn.font.italic): out.append(f"[op {s}] {ref} italic")
                if fcolor(cg.font.color)!=fcolor(cn.font.color): out.append(f"[op {s}] {ref} fontcolor {fcolor(cg.font.color)}!={fcolor(cn.font.color)}")
                if (cg.number_format or 'General')!=(cn.number_format or 'General'): out.append(f"[op {s}] {ref} numFmt {cg.number_format!r}!={cn.number_format!r}")
                ag=cg.alignment; an=cn.alignment
                if (ag.horizontal or None)!=(an.horizontal or None): out.append(f"[op {s}] {ref} align.h {ag.horizontal}!={an.horizontal}")
                if (ag.vertical or None)!=(an.vertical or None): out.append(f"[op {s}] {ref} align.v {ag.vertical}!={an.vertical}")
                if bool(ag.wrap_text)!=bool(an.wrap_text): out.append(f"[op {s}] {ref} wrap {ag.wrap_text}!={an.wrap_text}")

if __name__=="__main__":
    out=[]
    mode = sys.argv[1] if len(sys.argv)>1 else "all"
    if mode in ("all","inspect"): diff_inspect(out)
    if mode in ("all","op"): diff_openpyxl(out)
    print(f"TOTAL DIFFS: {len(out)}")
    for line in out: print(line)
