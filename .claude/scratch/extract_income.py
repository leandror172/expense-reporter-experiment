#!/usr/bin/env python3
"""WS-0b extractor: lift income (Receitas) records from 2023-2025 budget workbooks
into JSONL income logs.

Header-driven: month-start columns are read from the sheet's own header rows
(same layout as the expense extractor). Group boundaries are detected by "Total"
rows; the first sub-line label inside each group becomes the income_category.

Output schema (one JSON object per line):
  {
    "date":            "DD/MM/YYYY" | null,
    "value":           float,             # positive = income, negative = deduction
    "income_marker":   "income",          # fixed tag for routing (forward-compat with WS-C)
    "income_category": str,               # top-level group: "Salário", "Férias", "13° Salário", "Presente", "Outros"
    "income_label":    str,               # sub-line as written in the workbook
    "source":          str,               # source workbook filename (not the values!)
    "year":            int
  }

Personal values are NEVER hardcoded here. All numeric data is read from the workbook
at runtime and written only to the gitignored output file.
"""

import json, sys, datetime, os
from collections import Counter
import openpyxl

OLD = "/home/leandror/workspaces/expenses/old"
OUTDIR = os.path.join(OLD, "extracted")
INCOME_SHEET = "Receitas"

MONTHS = {
    "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4,
    "Maio": 5, "Junho": 6, "Julho": 7, "Agosto": 8,
    "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12
}
TOTALS = {"total", "totais", "total geral"}


def s(v):
    return "" if v is None else str(v).strip()


def is_total(v):
    return s(v).lower() in TOTALS


def fmt_date(raw, band_month, year):
    """Return DD/MM/YYYY, always. The band column is the authoritative MONTH and the
    source file the authoritative YEAR — both are always known for an emitted record.
    Only the DAY comes from the cell; when the Data cell is blank or unparseable (the
    common payslip case — monthly values recorded with no specific day), the day
    defaults to 01 so the month placement is never lost (WS-0b month-drop fix)."""
    day = 1
    if raw is not None and s(raw) != "":
        if isinstance(raw, (datetime.datetime, datetime.date)):
            day = raw.day
        else:
            try:
                parsed = int(s(raw).replace("-", "/").split("/")[0])
                if 1 <= parsed <= 31:
                    day = parsed
            except (ValueError, IndexError):
                day = 1
    return f"{day:02d}/{band_month:02d}/{year}"


def load_sheet_grid(ws):
    grid = []
    for row in ws.iter_rows(values_only=False):
        grid.append({c.column: c.value for c in row})
    return grid


# Canonical group names: first sub-line label -> canonical income_category.
# Applied when the first label in a block matches a known group-header pattern.
GROUP_CANONICAL = {
    "salário":          "Salário",
    "ferias normais":   "Férias",
    "férias normais":   "Férias",
    "13° integral":     "13° Salário",
    "presente":         "Presente",
    "outros":           "Outros",
    # Catch-all buckets used in 2023 for leisure sub-sheets (Jogos, Cigarro, etc.)
    "jogos":            "Outros",
    "cigarro":          "Outros",
    "café":             "Outros",
    "cerveja":          "Outros",
    "discos":           "Outros",
    "cinema/teatro":    "Outros",
    "restaurantes/bares": "Outros",
}


def normalize_group(first_label):
    """Map the first sub-line of a group block to a canonical income_category."""
    key = first_label.lower()
    for pattern, canonical in GROUP_CANONICAL.items():
        if key == pattern or key.startswith(pattern):
            return canonical
    return "Outros"


def parse_income_sheet(grid, year, fname):
    """Yield income record dicts. Header-driven, group-boundary detection via Total rows."""
    # Locate header rows (same logic as expense extractor)
    hm_idx = None
    for i, row in enumerate(grid):
        vals = {s(v) for v in row.values()}
        if vals & set(MONTHS) and hm_idx is None:
            hm_idx = i
    if hm_idx is None:
        return

    hm = grid[hm_idx]

    # Build band list: (month_num, item_col, date_col, value_col)
    bands = []
    for col, v in hm.items():
        m = MONTHS.get(s(v))
        if not m:
            continue
        # The Receitas sheet always has full Item/Data/Valor triplets
        bands.append((m, col, col + 1, col + 2))

    if not bands:
        return

    # Data rows start after the second header row (hm_idx + 1 = sub-header with Item/Data/Valor)
    data_start = hm_idx + 1 + 1   # skip "Item Data Valor" row
    label_col = 2  # col B (1-based) always holds sub-line labels in Receitas

    # Walk rows; track group via Total boundaries
    current_group = None
    first_in_group = True

    for row in grid[data_start:]:
        b_val = row.get(label_col)

        # Total row resets the group — Total appears in col C (item col of first band),
        # not always col B, so check all cells.
        if any(is_total(v) for v in row.values()):
            current_group = None
            first_in_group = True
            continue

        label = s(b_val)
        if not label:
            continue

        # First label after a group-reset sets the group name
        if first_in_group:
            current_group = normalize_group(label)
            first_in_group = False

        # Emit one record per band that has a numeric value
        for (month, item_col, date_col, value_col) in bands:
            value = row.get(value_col)
            if not isinstance(value, (int, float)) or value == 0:
                continue
            date_raw = row.get(date_col)
            item_raw = row.get(item_col)
            date_str = fmt_date(date_raw, month, year)
            # item column in Receitas sometimes holds a free-text note; keep it
            item_note = s(item_raw) if item_raw is not None else ""
            yield {
                "date":            date_str,
                "value":           float(value),
                "income_marker":   "income",
                "income_category": current_group or "Outros",
                "income_label":    label,
                "item_note":       item_note,   # free-text note from Item col (may be "")
                "source":          fname,
                "year":            year,
            }


def main():
    os.makedirs(OUTDIR, exist_ok=True)

    files = {
        # 2022 workbook has no Receitas sheet — skipped intentionally
        2023: "Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2023.xlsx",
        2024: "Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2024.xlsx",
        2025: "Planilha_Normalized_Final_copy.xlsx",
    }

    all_records = []
    year_counts = Counter()

    for year, fname in files.items():
        path = os.path.join(OLD, fname)
        wb = openpyxl.load_workbook(path, data_only=True)
        if INCOME_SHEET not in wb.sheetnames:
            print(f"{year}: no Receitas sheet — skipped")
            continue

        grid = load_sheet_grid(wb[INCOME_SHEET])
        year_recs = list(parse_income_sheet(grid, year, fname))
        all_records.extend(year_recs)
        year_counts[year] = len(year_recs)
        print(f"{year}: {len(year_recs)} income records")

    outfile = os.path.join(OUTDIR, "income_log.jsonl")
    with open(outfile, "w", encoding="utf-8") as f:
        for rec in all_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    print(f"\nTotal: {sum(year_counts.values())} records → {outfile}")
    print("\nBreakdown by year+category:")
    cat_counts = Counter()
    for rec in all_records:
        cat_counts[(rec["year"], rec["income_category"])] += 1
    for (yr, cat), n in sorted(cat_counts.items()):
        print(f"  {yr}  {cat:<20}  {n:>3} records")


if __name__ == "__main__":
    main()
