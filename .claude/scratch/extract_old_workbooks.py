#!/usr/bin/env python3
"""One-off (5.R4) extractor: lift expense records from the 2022-2024 budget
workbooks into entries-log JSONL ({item,date,value,type,category,subcategory}).

Header-driven: month-start columns + Item/Data/Valor triplets are read from each
sheet's own header rows, so it handles both the standard layout (2023/24) and the
2022 layout (shifted cat/sub cols, value-only Jan-Oct bands).

Faithful extraction: records carry the source's own category/subcategory labels.
A separate conformance report flags tuples that don't match config/taxonomy.json.
No personal values are printed; outputs go OUTSIDE the repo.
"""
import json, sys, datetime, os
from collections import Counter, defaultdict
import openpyxl
from openpyxl.utils import get_column_letter

OLD = "/home/leandror/workspaces/expenses/old"
OUTDIR = os.path.join(OLD, "extracted")
TAXONOMY = "/mnt/i/workspaces/expenses/code/expense-reporter/config/taxonomy.json"
EXPENSE_SHEETS = ["Fixas", "Variáveis", "Extras", "Adicionais"]
MONTHS = {"Janeiro":1,"Fevereiro":2,"Março":3,"Abril":4,"Maio":5,"Junho":6,
          "Julho":7,"Agosto":8,"Setembro":9,"Outubro":10,"Novembro":11,"Dezembro":12}
TOTALS = {"total", "totais", "total geral"}

# Alias map: faithful source (type,category,subcategory) -> taxonomy-conforming target.
# Decided 2026-06-20 (see project_workbook_extraction_5r4 memory). The map itself lives in
# a gitignored JSON (taxonomy-label fragments) so this script stays clean for git history.
ALIASES_FILE = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "classification", "extraction-aliases.json"
)

def load_aliases(path):
    """Load the gitignored alias map: JSON list of {source, target} -> {tuple: tuple}."""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return {tuple(a["source"]): tuple(a["target"]) for a in data["aliases"]}

ALIASES = load_aliases(ALIASES_FILE)

def s(v):
    return "" if v is None else str(v).strip()

def is_total(v):
    return s(v).lower() in TOTALS

def fmt_date(raw, band_month, year):
    """Return DD/MM/YYYY or None. The band column is the authoritative MONTH and the
    source file the authoritative YEAR; only the DAY is taken from the cell. This
    discards Excel's auto-year/auto-month artifacts (bare 'd/m' typed cells get stored
    as real datetimes stamped with the workbook's edit-year)."""
    if raw is None or s(raw) == "":
        return None
    if isinstance(raw, (datetime.datetime, datetime.date)):
        day = raw.day
    else:
        try:
            day = int(s(raw).replace("-", "/").split("/")[0])
        except (ValueError, IndexError):
            return None
    if not 1 <= day <= 31:
        return None
    return f"{day:02d}/{band_month:02d}/{year}"

def load_sheet_grid(ws):
    """Return list of rows; each row is a dict {col_index(1-based): value}."""
    grid = []
    for row in ws.iter_rows(values_only=False):
        grid.append({c.column: c.value for c in row})
    return grid

def parse_sheet(grid, sheet_name, year):
    """Yield (record_dict, is_partial). Header-driven band geometry."""
    # locate header rows
    hc_idx = hm_idx = None
    for i, row in enumerate(grid):
        vals = {s(v) for v in row.values()}
        if "Item" in vals and hc_idx is None:
            hc_idx = i
        if vals & set(MONTHS) and hm_idx is None:
            hm_idx = i
    if hc_idx is None or hm_idx is None:
        return
    hc = grid[hc_idx]
    hm = grid[hm_idx]

    # bands: month-start col -> (month_num, item_col|None, date_col|None, value_col)
    bands = []
    for col, v in hm.items():
        m = MONTHS.get(s(v))
        if not m:
            continue
        if s(hc.get(col)) == "Item":               # full band
            bands.append((m, col, col + 1, col + 2))
        else:                                        # value-only band (2022 Jan-Oct)
            bands.append((m, None, None, col))
    if not bands:
        return
    first_band_col = min(b[3] if b[1] is None else b[1] for b in bands)

    # detect cat/sub label columns: the two label cols (left of first band) that
    # carry the most text across data rows.
    label_cols = [c for c in range(1, first_band_col)]
    text_count = Counter()
    for row in grid[hc_idx + 1:]:
        for c in label_cols:
            val = row.get(c)
            if isinstance(val, str) and s(val) and not is_total(val):
                text_count[c] += 1
    ranked = sorted(text_count, key=lambda c: (-text_count[c], c))
    chosen = sorted(ranked[:2])
    if len(chosen) < 2:
        return
    cat_col, sub_col = chosen[0], chosen[1]

    cur_cat = cur_sub = None
    for row in grid[hc_idx + 1:]:
        if any(is_total(v) for v in row.values()):
            continue  # total row
        cval = row.get(cat_col)
        if isinstance(cval, str) and s(cval):
            cur_cat = s(cval)
        sval = row.get(sub_col)
        if isinstance(sval, str) and s(sval):
            cur_sub = s(sval)
        if not cur_sub:
            continue
        for (month, item_col, date_col, value_col) in bands:
            value = row.get(value_col)
            if not isinstance(value, (int, float)) or value == 0:
                continue
            item = s(row.get(item_col)) if item_col else ""
            if is_total(item):
                continue
            date = fmt_date(row.get(date_col), month, year) if date_col else None
            rec = {"item": item, "date": date, "value": float(value),
                   "type": sheet_name, "category": cur_cat, "subcategory": cur_sub}
            yield rec, (item_col is None)

def main():
    os.makedirs(OUTDIR, exist_ok=True)
    tax = json.load(open(TAXONOMY))
    valid_paths = set()
    for ty in tax["types"]:
        for c in ty["categories"]:
            for sub in c["subcategories"]:
                valid_paths.add((ty["name"], c["name"], sub))

    files = {2022: "Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2022.xlsx",
             2023: "Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2023.xlsx",
             2024: "Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2024.xlsx",
             2025: "Planilha_Normalized_Final_copy.xlsx"}

    nonconforming = Counter()
    grand_full = grand_partial = 0
    for year, fname in files.items():
        wb = openpyxl.load_workbook(os.path.join(OLD, fname), data_only=True)
        full, partial = [], []
        for sheet in EXPENSE_SHEETS:
            if sheet not in wb.sheetnames:
                continue
            grid = load_sheet_grid(wb[sheet])
            for rec, is_partial in parse_sheet(grid, sheet, year):
                tup = (rec["type"], rec["category"], rec["subcategory"])
                if tup not in valid_paths:
                    nonconforming[tup] += 1
                (partial if is_partial else full).append(rec)
        with open(os.path.join(OUTDIR, f"extracted-{year}.jsonl"), "w") as f:
            for r in full:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
        # normalized copy: apply alias remap to taxonomy-conforming paths
        with open(os.path.join(OUTDIR, f"normalized-{year}.jsonl"), "w") as f:
            for r in full:
                tup = (r["type"], r["category"], r["subcategory"])
                if tup in ALIASES:
                    r = dict(r)
                    r["type"], r["category"], r["subcategory"] = ALIASES[tup]
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
        if partial:
            with open(os.path.join(OUTDIR, f"partial-{year}.jsonl"), "w") as f:
                for r in partial:
                    f.write(json.dumps(r, ensure_ascii=False) + "\n")
        grand_full += len(full); grand_partial += len(partial)
        print(f"{year}: {len(full)} full records, {len(partial)} value-only (partial)")

    print(f"\nTOTAL: {grand_full} full, {grand_partial} partial")
    print("\n=== Non-conforming (type, category, subcategory) tuples vs taxonomy ===")
    for tup, n in sorted(nonconforming.items(), key=lambda kv: -kv[1]):
        print(f"  {n:>4}  {tup}")

if __name__ == "__main__":
    main()
