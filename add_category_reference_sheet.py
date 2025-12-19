import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

# Load the workbook
wb = openpyxl.load_workbook('Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2025.xlsx')

# Get the reference sheet
ws_ref = wb['Listas de itens']

# Structure to hold categories and subcategories
categories = {}

# Define the structure based on the analysis
# Main categories: Receitas, Investimentos, Fixas, Variáveis, Extras, Adicionais

# Receitas (Income)
categories['Receitas'] = [
    'Salário',
    'Ajuda de custo',
    'Aluguel',
    'Pensão',
    'Horas extras',
    '13º salário',
    'Férias',
    'Outros'
]

# Investimentos (Investments)
categories['Investimentos'] = [
    'Ações',
    'Tesouro Direto',
    'Renda fixa',
    'Previdência privada',
    'Outros'
]

# Fixas (Fixed Expenses)
categories['Fixas'] = {
    'Habitação': [
        'Aluguel',
        'Condomínio',
        'Prestação da casa',
        'Seguro da casa',
        'Celular Leandro',
        'Celular Ana',
        'Diarista',
        'Google storage',
        'Internet'
    ],
    'Lazer': [
        'Netflix',
        'Amazon',
        'HBO',
        'Spotify',
        'Globoplay',
        'Clube do Malte',
        'Gamepass'
    ],
    'Transporte': [
        'Prestação do carro',
        'Seguro do carro',
        'Estacionamento'
    ],
    'Saúde': [
        'Seguro saúde',
        'Plano odontológico',
        'Pilates',
        'Laya',
        'Santa Cannabis',
        'Plano de saúde'
    ],
    'Educação': [
        'Colégio',
        'Faculdade',
        'Bateria'
    ],
    'Pet': [
        'Orion - Óleo',
        'Orion - Plano',
        'Lilly - Plano',
        'Ambos'
    ],
    'Impostos': [
        'IPTU',
        'Licenciamento',
        'INSS',
        'IRFF',
        'IPVA'
    ],
    'Outros': [
        'alguma coisa sindicato',
        'Contribuição sindicato',
        'Apoia-se 4i20',
        'Seguro de vida'
    ]
}

# Variáveis (Variable Expenses)
categories['Variáveis'] = {
    'Habitação': [
        'Luz',
        'Água',
        'Produtos',
        'Gás'
    ],
    'Impostos': [
        'IOF'
    ],
    'Transporte': [
        'Metrô',
        'Ônibus',
        'Uber/Taxi',
        'Combustível',
        'Pedágio',
        'Estacionamento'
    ],
    'Alimentação / Limpeza': [
        'Supermercado VA',
        'Supermercado',
        'Feira',
        'Açougue',
        'Padaria'
    ],
    'Pet': [
        'Orion - Consultas',
        'Orion - Ração',
        'Orion - Banho',
        'Lilly - Consultas',
        'Lilly - Ração úmida',
        'Lilly - Banho',
        'Lilly - Ração',
        'Ambos - Ração',
        'Ambos - Pet shop',
        'Ambos - Petiscos'
    ],
    'Saúde': [
        'Exames',
        'Consultas',
        'Acupuntura / Massagem',
        'Óleo/flor cannabis',
        'Dentista',
        'Farmácia'
    ],
    'Anita': [
        'Terpenos',
        'Ingredientes',
        'Ingrediente chocolate',
        'Empréstimo'
    ],
    'Cuidados pessoais': [
        'Cabeleireiro',
        'Manicure',
        'Produtos',
        'Academia'
    ]
}

# Extras (Extra Expenses)
categories['Extras'] = {
    'Saúde': [
        'Médico',
        'Dentista',
        'Hospital'
    ],
    'Manutenção/ prevenção': [
        'Carro',
        'Casa',
        'Mudança'
    ],
    'Pets': [
        'Orion - Vacinas',
        'Orion - outros',
        'Lilly - Vacinas',
        'Lilly - outros',
        'Ambos'
    ],
    'Advogado': [
        'HC grow'
    ],
    'Educação': [
        'Material escolar',
        'Rematrícula faculdade',
        'Uniforme'
    ]
}

# Adicionais (Additional Expenses)
categories['Adicionais'] = {
    'Lazer': [
        'Viagens',
        'Grow',
        'Jogos',
        'Computador ($)',
        'Cigarro',
        'Café',
        'Cerveja',
        'Livros',
        'Cinema/teatro'
    ]
}

# Create a new worksheet for the category reference
if 'Referência de Categorias' in wb.sheetnames:
    del wb['Referência de Categorias']

ws_new = wb.create_sheet('Referência de Categorias')

# Set up header formatting
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
subheader_font = Font(bold=True, color='FFFFFF', size=11)
category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
category_font = Font(bold=True, size=10)

# Write title
ws_new['A1'] = 'REFERÊNCIA DE CATEGORIAS E SUB-CATEGORIAS'
ws_new['A1'].font = Font(bold=True, size=14)
ws_new['A1'].alignment = Alignment(horizontal='center')
ws_new.merge_cells('A1:C1')

# Add description
ws_new['A2'] = 'Esta planilha lista todas as categorias e sub-categorias para referência em código/macros'
ws_new['A2'].font = Font(italic=True, size=10)
ws_new.merge_cells('A2:C2')

row = 4

# Write headers
ws_new['A' + str(row)] = 'Tipo Principal'
ws_new['B' + str(row)] = 'Categoria'
ws_new['C' + str(row)] = 'Sub-categoria'
for col in ['A', 'B', 'C']:
    cell = ws_new[col + str(row)]
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

row += 1

# Write Receitas
start_row = row
for subcategory in categories['Receitas']:
    ws_new['A' + str(row)] = 'Receitas'
    ws_new['C' + str(row)] = subcategory
    ws_new['A' + str(row)].fill = category_fill
    ws_new['A' + str(row)].font = category_font
    row += 1

# Write Investimentos
start_row = row
for subcategory in categories['Investimentos']:
    ws_new['A' + str(row)] = 'Investimentos'
    ws_new['C' + str(row)] = subcategory
    ws_new['A' + str(row)].fill = category_fill
    ws_new['A' + str(row)].font = category_font
    row += 1

# Write Fixas (with categories)
for category, subcategories in categories['Fixas'].items():
    start_row = row
    for subcategory in subcategories:
        ws_new['A' + str(row)] = 'Fixas'
        ws_new['B' + str(row)] = category
        ws_new['C' + str(row)] = subcategory
        if row == start_row:
            ws_new['A' + str(row)].fill = category_fill
            ws_new['A' + str(row)].font = category_font
            ws_new['B' + str(row)].fill = subheader_fill
            ws_new['B' + str(row)].font = subheader_font
        row += 1

# Write Variáveis (with categories)
for category, subcategories in categories['Variáveis'].items():
    start_row = row
    for subcategory in subcategories:
        ws_new['A' + str(row)] = 'Variáveis'
        ws_new['B' + str(row)] = category
        ws_new['C' + str(row)] = subcategory
        if row == start_row:
            ws_new['A' + str(row)].fill = category_fill
            ws_new['A' + str(row)].font = category_font
            ws_new['B' + str(row)].fill = subheader_fill
            ws_new['B' + str(row)].font = subheader_font
        row += 1

# Write Extras (with categories)
for category, subcategories in categories['Extras'].items():
    start_row = row
    for subcategory in subcategories:
        ws_new['A' + str(row)] = 'Extras'
        ws_new['B' + str(row)] = category
        ws_new['C' + str(row)] = subcategory
        if row == start_row:
            ws_new['A' + str(row)].fill = category_fill
            ws_new['A' + str(row)].font = category_font
            ws_new['B' + str(row)].fill = subheader_fill
            ws_new['B' + str(row)].font = subheader_font
        row += 1

# Write Adicionais (with categories)
for category, subcategories in categories['Adicionais'].items():
    start_row = row
    for subcategory in subcategories:
        ws_new['A' + str(row)] = 'Adicionais'
        ws_new['B' + str(row)] = category
        ws_new['C' + str(row)] = subcategory
        if row == start_row:
            ws_new['A' + str(row)].fill = category_fill
            ws_new['A' + str(row)].font = category_font
            ws_new['B' + str(row)].fill = subheader_fill
            ws_new['B' + str(row)].font = subheader_font
        row += 1

# Adjust column widths
ws_new.column_dimensions['A'].width = 20
ws_new.column_dimensions['B'].width = 25
ws_new.column_dimensions['C'].width = 30

# Add borders to all cells with data
from openpyxl.styles import Border, Side
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row_cells in ws_new.iter_rows(min_row=4, max_row=row-1, min_col=1, max_col=3):
    for cell in row_cells:
        cell.border = thin_border
        if cell.column == 3:  # Sub-category column
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Save the workbook
wb.save('Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2025.xlsx')

print("Nova planilha 'Referencia de Categorias' adicionada com sucesso!")
print(f"Total de linhas criadas: {row - 5}")
print(f"Estrutura:")
print(f"  - Receitas: {len(categories['Receitas'])} itens")
print(f"  - Investimentos: {len(categories['Investimentos'])} itens")
print(f"  - Fixas: {sum(len(v) for v in categories['Fixas'].values())} itens em {len(categories['Fixas'])} categorias")
print(f"  - Variaveis: {sum(len(v) for v in categories['Variáveis'].values())} itens em {len(categories['Variáveis'])} categorias")
print(f"  - Extras: {sum(len(v) for v in categories['Extras'].values())} itens em {len(categories['Extras'])} categorias")
print(f"  - Adicionais: {sum(len(v) for v in categories['Adicionais'].values())} itens em {len(categories['Adicionais'])} categorias")
