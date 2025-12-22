"""
Script to standardize column layout across all expense sheets.
Currently: Fixas/Extras/Adicionais start at C, Variáveis starts at D
After: All sheets will start at D (by inserting column C in the 3 sheets)
"""

import openpyxl
from copy import copy

print("Loading workbook...")
wb = openpyxl.load_workbook('Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2025.xlsx')

# Sheets that currently start at column C (need an extra column)
sheets_to_fix = ['Fixas', 'Extras', 'Adicionais']

print("\nStandardizing column layout to start at D for all sheets...")
print("=" * 80)

for sheet_name in sheets_to_fix:
    ws = wb[sheet_name]

    print(f"\n{sheet_name}:")
    print(f"  Current: Month columns start at C")
    print(f"  Action: Inserting blank column C, shifting months to D")

    # Insert a new column at position C (column 3)
    # This shifts everything from C onwards to the right
    ws.insert_cols(3)  # Insert at column 3 (C)

    print(f"  Result: Month columns now start at D")

    # Verify: Check row 1 for month names
    print(f"  Verification:")
    for col_idx in range(1, 10):
        cell_value = ws.cell(1, col_idx).value
        if cell_value and isinstance(cell_value, str):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"    Column {col_letter}: {cell_value}")

print("\n" + "=" * 80)
print("Column standardization complete!")
print("\nAll sheets now have:")
print("  - Column A: Category")
print("  - Column B: Subcategory")
print("  - Column C: (blank/spacing)")
print("  - Column D onwards: Month data (Janeiro starts at D)")
print("\nThis eliminates the need for sheet-specific column logic in Go code!")

# Save workbook
print("\nSaving workbook...")
wb.save('Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2025.xlsx')

print("\n" + "=" * 80)
print("SUCCESS!")
print("=" * 80)
print("All expense sheets now use consistent column layout")
print("Go code can use single column mapping for all sheets")
