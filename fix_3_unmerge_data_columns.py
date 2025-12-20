"""
Script to unmerge cells in data entry columns (D onwards) while preserving header/label merges.
This prevents issues with cell addressing when inserting expense data.
"""

import openpyxl
from copy import copy

print("Loading workbook...")
wb = openpyxl.load_workbook('Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2025.xlsx')

expense_sheets = ['Fixas', 'Variáveis', 'Extras', 'Adicionais']

print("\nAnalyzing and unmerging data entry columns...")
print("=" * 80)

total_unmerged = 0

for sheet_name in expense_sheets:
    ws = wb[sheet_name]

    print(f"\n{sheet_name}:")
    print(f"  Total merged ranges before: {len(ws.merged_cells.ranges)}")

    # Get all merged cell ranges (make a copy of the list)
    merged_ranges = list(ws.merged_cells.ranges)

    unmerged_count = 0

    for merged_range in merged_ranges:
        # Get the range coordinates
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        min_row = merged_range.min_row
        max_row = merged_range.max_row

        # Unmerge if it includes or is entirely in data columns (D = 4 onwards)
        # But preserve merges entirely in columns A-C (1-3)
        if max_col >= 4:  # Touches or is in data columns
            # Save the value and style from the top-left cell before unmerging
            top_left_cell = ws.cell(min_row, min_col)
            saved_value = top_left_cell.value
            saved_style = copy(top_left_cell._style)

            # Unmerge
            ws.unmerge_cells(str(merged_range))
            unmerged_count += 1

            # Restore value and style to top-left cell
            top_left_cell.value = saved_value
            top_left_cell._style = saved_style

            print(f"  Unmerged: {merged_range} (touched data columns)")

    print(f"  Unmerged: {unmerged_count} ranges")
    print(f"  Merged ranges remaining: {len(ws.merged_cells.ranges)} (header/label area)")

    total_unmerged += unmerged_count

print("\n" + "=" * 80)
print(f"Total unmerged across all sheets: {total_unmerged}")

# Save workbook
print("\nSaving workbook...")
wb.save('Planilha_BMeFBovespa_Leandro_OrcamentoPessoal-2025.xlsx')

print("\n" + "=" * 80)
print("SUCCESS!")
print("=" * 80)
print("Data entry columns (D onwards) are now unmerged")
print("Header and label merges (columns A-C) preserved")
print("This will prevent cell addressing issues during data insertion")
